#!/usr/bin/env python3
"""
项目跟踪系统 v3.0 - Flask + Jinja2
纯服务器端渲染，支持两级Tab导航、拖拽排序、会议管理
"""

import os
import json
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file
import html
import io

app = Flask(__name__)
app.config['JSON_AS_cii'] = False

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
PROJECTS_FILE = os.path.join(DATA_DIR, 'projects.json')

# ===== 服务器代码更新检测 =====
SERVER_FILE = os.path.abspath(__file__)
_server_mtime = os.path.getmtime(SERVER_FILE)
_restart_warned = False


@app.after_request
def set_restart_header(response):
    """检测 server.py 是否被修改，每次响应都追加提示头"""
    global _server_mtime
    if os.path.getmtime(SERVER_FILE) > _server_mtime:
        response.headers['X-Server-Restart-Required'] = 'true'
        response.headers['X-Restart-Message'] = 'server.py 已更新，请重启服务器'
    return response


def load_data():
    os.makedirs(DATA_DIR, exist_ok=True)
    if os.path.exists(PROJECTS_FILE):
        with open(PROJECTS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {'四建': [], '亚太': [], 'meetings': []}


def save_data(data):
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(PROJECTS_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def get_valid_tabs(data):
    """动态获取所有合法的tab值"""
    return ['all'] + [g for g in data.keys() if g != 'meetings']


def valid_tab(tab, data):
    """检查tab是否合法，不合法则返回'all'"""
    return tab if tab in get_valid_tabs(data) else 'all'


def sort_projects_by_order(projects):
    """按 sort_order 排序，未设置则按创建顺序"""
    return sorted(projects, key=lambda p: p.get('sort_order', 999))


# ==================== 两级Tab路由 ====================

@app.route('/')
def index():
    data = load_data()
    section = request.args.get('section', 'projects')
    tab = request.args.get('tab', 'all')

    # meetings section 单独处理
    if section == 'meetings':
        return render_template('index.html',
                               projects=data,
                               active_section='meetings',
                               active_tab=tab,
                               stats=None)

    # 每日任务跳转到 daily.html
    if section == 'daily':
        return redirect('/daily.html')

    # 默认 section=projects
    if section not in ('projects', ''):
        section = 'projects'

    # 计算全部项目的统计数据（仅在 all tab 显示）
    stats = {'total_proj': 0, 'active_proj': 0, 'pending_proj': 0,
             'total_tasks': 0, 'done_tasks': 0, 'task_rate': 0}
    for group, projects_list in data.items():
        if group == 'meetings':
            continue
        for p in projects_list:
            stats['total_proj'] += 1
            status_str = p.get('status_text') or p.get('status') or ''
            if '完成' not in status_str:
                stats['active_proj'] += 1
            if '待' in status_str or '评估' in status_str:
                stats['pending_proj'] += 1
            for t in p.get('tasks', []):
                stats['total_tasks'] += 1
                if t.get('done'):
                    stats['done_tasks'] += 1
    if stats['total_tasks'] > 0:
        stats['task_rate'] = round(stats['done_tasks'] / stats['total_tasks'] * 100, 1)

    return render_template('index.html',
                           projects=data,
                           active_section='projects',
                           active_tab=tab,
                           stats=stats)


# ==================== 项目拖拽排序 API ====================

@app.route('/api/reorder_projects', methods=['POST'])
def reorder_projects():
    """更新项目的 sort_order"""
    req = request.get_json()
    group = req.get('group')
    project_ids = req.get('order', [])  # 新的排序ID列表 [id1, id2, ...]

    data = load_data()
    if not group or group not in data:
        return jsonify({'error': 'Invalid group'}), 400

    # 更新 sort_order
    for idx, pid in enumerate(project_ids):
        for p in data[group]:
            if p.get('id') == pid:
                p['sort_order'] = idx
                break

    save_data(data)
    return jsonify({'success': True})


# ==================== 会议管理 API ====================

@app.route('/api/meetings', methods=['GET'])
def get_meetings():
    """获取所有会议"""
    data = load_data()
    meetings = data.get('meetings', [])
    # 按日期排序，未完成在前
    today = datetime.now().strftime('%Y-%m-%d')
    upcoming = []
    past = []
    for m in meetings:
        m_date = m.get('date', '')
        # 解析日期（支持格式：周二 04-02 或 2026-04-02）
        if m.get('completed'):
            past.append(m)
        elif '-' in str(m_date) and len(str(m_date)) >= 10:
            # 完整日期格式
            if m_date < today:
                past.append(m)
            else:
                upcoming.append(m)
        elif '-' in str(m_date) and len(str(m_date).split('-')[-1]) <= 2:
            # 短格式如 周二 04-02，算作未来
            upcoming.append(m)
        else:
            upcoming.append(m)

    upcoming.sort(key=lambda x: x.get('date', ''))
    past.sort(key=lambda x: x.get('date', ''), reverse=True)
    return jsonify({'meetings': upcoming + past})


@app.route('/api/meetings', methods=['POST'])
def create_meeting():
    """新建会议"""
    req = request.get_json()
    name = req.get('name', '').strip()
    date = req.get('date', '').strip()
    time = req.get('time', '').strip()
    remark = req.get('remark', '').strip()

    if not name:
        return jsonify({'error': '会议名称不能为空'}), 400

    data = load_data()
    meetings = data.get('meetings', [])

    # 生成ID
    existing_ids = [m.get('id', '') for m in meetings if m.get('id', '').startswith('meet-')]
    nums = [int(x.split('-')[-1]) for x in existing_ids if x.split('-')[-1].isdigit()]
    next_num = max(nums) + 1 if nums else 1
    meeting_id = f'meet-{str(next_num).zfill(3)}'

    new_meeting = {
        'id': meeting_id,
        'name': html.escape(name),
        'date': date,
        'time': time,
        'remark': html.escape(remark),
        'completed': False,
        'createdAt': datetime.now().strftime('%Y-%m-%d')
    }
    meetings.append(new_meeting)
    data['meetings'] = meetings
    save_data(data)
    return jsonify({'meeting': new_meeting})


@app.route('/api/meetings/<meeting_id>', methods=['DELETE'])
def delete_meeting(meeting_id):
    """删除会议"""
    data = load_data()
    meetings = data.get('meetings', [])
    data['meetings'] = [m for m in meetings if m.get('id') != meeting_id]
    save_data(data)
    return jsonify({'success': True})


@app.route('/api/meetings/<meeting_id>/toggle', methods=['POST'])
def toggle_meeting(meeting_id):
    """标记完成/未完成"""
    data = load_data()
    meetings = data.get('meetings', [])
    for m in meetings:
        if m.get('id') == meeting_id:
            m['completed'] = not m.get('completed', False)
            break
    save_data(data)
    return jsonify({'success': True})


@app.route('/api/meetings/<meeting_id>', methods=['PUT'])
def update_meeting(meeting_id):
    """更新会议"""
    req = request.get_json()
    data = load_data()
    meetings = data.get('meetings', [])
    for m in meetings:
        if m.get('id') == meeting_id:
            if 'name' in req:
                m['name'] = html.escape(req['name'].strip())
            if 'date' in req:
                m['date'] = req['date'].strip()
            if 'time' in req:
                m['time'] = req['time'].strip()
            if 'remark' in req:
                m['remark'] = html.escape(req['remark'].strip())
            break
    save_data(data)
    return jsonify({'success': True})


# ==================== 项目任务相关路由（保持兼容）====================

@app.route('/toggle_task', methods=['POST'])
def toggle_task():
    data = load_data()
    group = request.form.get('group', '')
    project_id = request.form.get('project_id', '')
    task_text = request.form.get('task_text', '')
    tab = request.form.get('tab', 'all')
    section = request.form.get('section', 'projects')

    for p in data.get(group, []):
        if p.get('id') == project_id:
            for t in p.get('tasks', []):
                if t.get('text') == task_text:
                    t['done'] = not t.get('done', False)
                    break
            break

    save_data(data)
    return redirect(url_for('index', section=section, tab=valid_tab(tab, data)) + f"#project-{project_id}")


@app.route('/update_priority', methods=['POST'])
def update_priority():
    data = load_data()
    group = request.form.get('group', '')
    project_id = request.form.get('project_id', '')
    task_text = request.form.get('task_text', '')
    tab = request.form.get('tab', 'all')
    section = request.form.get('section', 'projects')

    for p in data.get(group, []):
        if p.get('id') == project_id:
            for t in p.get('tasks', []):
                if t.get('text') == task_text:
                    priorities = ['low', 'medium', 'high']
                    current = t.get('priority', 'medium')
                    idx = priorities.index(current)
                    t['priority'] = priorities[(idx + 1) % 3]
                    break
            break

    save_data(data)
    return redirect(url_for('index', section=section, tab=valid_tab(tab, data)) + f"#project-{project_id}")


@app.route('/update_due', methods=['POST'])
def update_due():
    data = load_data()
    group = request.form.get('group', '')
    project_id = request.form.get('project_id', '')
    task_text = request.form.get('task_text', '')
    new_due = request.form.get('due', '')
    tab = request.form.get('tab', 'all')
    section = request.form.get('section', 'projects')

    for p in data.get(group, []):
        if p.get('id') == project_id:
            for t in p.get('tasks', []):
                if t.get('text') == task_text:
                    t['due'] = new_due
                    break
            break

    save_data(data)
    anchor = f"project-{project_id}"
    return redirect(url_for('index', section=section, tab=valid_tab(tab, data)) + f"#{anchor}")


@app.route('/clear_due', methods=['POST'])
def clear_due():
    data = load_data()
    group = request.form.get('group', '')
    project_id = request.form.get('project_id', '')
    task_text = request.form.get('task_text', '')
    tab = request.form.get('tab', 'all')
    section = request.form.get('section', 'projects')

    for p in data.get(group, []):
        if p.get('id') == project_id:
            for t in p.get('tasks', []):
                if t.get('text') == task_text:
                    t['due'] = ''
                    break
            break

    save_data(data)
    return redirect(url_for('index', section=section, tab=valid_tab(tab, data)) + f"#project-{project_id}")


@app.route('/update_project', methods=['POST'])
def update_project():
    data = load_data()
    group = request.form.get('group', '')
    project_id = request.form.get('project_id', '')
    tab = request.form.get('tab', 'all')
    section = request.form.get('section', 'projects')

    for i, p in enumerate(data.get(group, [])):
        if p.get('id') == project_id:
            p['name'] = request.form.get('name', p.get('name', ''))
            p['status_text'] = request.form.get('status_text', '')
            p['status_color'] = request.form.get('status_color', '')
            p['node'] = request.form.get('node', p.get('node', ''))
            p['deadline'] = request.form.get('deadline', p.get('deadline', ''))
            p['issues'] = request.form.get('issues', p.get('issues', ''))

            tasks_text = request.form.get('tasks', '')
            tasks = []
            for line in tasks_text.split('\n'):
                line = line.strip()
                if line:
                    done = line.startswith('✅')
                    text = line.replace('✅', '').replace('☐', '').strip()
                    tasks.append({'text': text, 'done': done, 'priority': 'medium', 'due': ''})
            p['tasks'] = tasks
            break

    save_data(data)
    return redirect(url_for('index', section=section, tab=valid_tab(tab, data)))


@app.route('/daily.html')
def daily():
    return render_template('daily.html')


@app.route('/update_task_name', methods=['POST'])
def update_task_name():
    data = load_data()
    group = request.form.get('group', '')
    project_id = request.form.get('project_id', '')
    old_text = request.form.get('old_text', '')
    new_text = request.form.get('new_text', '')
    tab = request.form.get('tab', 'all')
    section = request.form.get('section', 'projects')

    for p in data.get(group, []):
        if p.get('id') == project_id:
            for t in p.get('tasks', []):
                if t.get('text') == old_text:
                    t['text'] = new_text
                    break
            break

    save_data(data)
    return redirect(url_for('index', section=section, tab=valid_tab(tab, data)) + f"#project-{project_id}")


@app.route('/update_task_detail', methods=['POST'])
def update_task_detail():
    """更新任务的完整详情：任务名称、优先级、截止日期、意见备注"""
    data = load_data()
    group = request.form.get('group', '')
    project_id = request.form.get('project_id', '')
    task_text = request.form.get('task_text', '')
    new_text = request.form.get('text', '')
    priority = request.form.get('priority', 'medium')
    due = request.form.get('due', '')
    opinion = request.form.get('opinion', '')
    tab = request.form.get('tab', 'all')
    section = request.form.get('section', 'projects')

    for p in data.get(group, []):
        if p.get('id') == project_id:
            for t in p.get('tasks', []):
                if t.get('text') == task_text:
                    t['text'] = new_text
                    t['priority'] = priority
                    t['due'] = due
                    t['opinion'] = opinion
                    break
            break

    save_data(data)
    return redirect(url_for('index', section=section, tab=valid_tab(tab, data)) + f"#project-{project_id}")


# ==================== 每日任务模块 API ====================

def json_response(data):
    return jsonify(data)

TASK_STATUSES = ['pending', 'in_progress', 'completed']
TASK_SOURCES = ['planned', 'temp']
TASK_PRIORITIES = ['low', 'medium', 'high']


@app.route('/api/daily_tasks', methods=['GET'])
def get_daily_tasks():
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    with open('data/daily_tasks.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    tasks = data.get('tasks', [])
    if date_from:
        tasks = [t for t in tasks if t.get('date', '') >= date_from]
    if date_to:
        tasks = [t for t in tasks if t.get('date', '') <= date_to]
    tasks.sort(key=lambda t: t.get('date', ''), reverse=True)
    return json_response({'tasks': tasks})


@app.route('/api/daily_tasks', methods=['POST'])
def create_daily_task():
    req = request.get_json()
    # 输入校验
    if not req.get('content', '').strip():
        return jsonify({'error': '任务内容不能为空'}), 400
    date_val = req.get('date', '').strip()
    if date_val:
        try:
            datetime.strptime(date_val, '%Y-%m-%d')
        except ValueError:
            return jsonify({'error': '日期格式错误，应为YYYY-MM-DD'}), 400
    with open('data/daily_tasks.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    data['last_id'] += 1
    new_task = {
        'id': data['last_id'],
        'content': html.escape(req.get('content', '')),
        'date': req.get('date', ''),
        'status': 'pending',
        'source': req.get('source', 'planned'),
        'source_dept': req.get('source_dept', ''),
        'priority': req.get('priority', 'medium'),
        'notes': html.escape(req.get('notes', '')),
        'tags': req.get('tags', []),
        'project': req.get('project', ''),
        'created_at': str(datetime.now()),
        'completed_at': None
    }
    data['tasks'].append(new_task)
    with open('data/daily_tasks.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return json_response({'task': new_task})


@app.route('/api/daily_tasks/<int:task_id>', methods=['PUT'])
def update_daily_task(task_id):
    req = request.get_json()
    with open('data/daily_tasks.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    task_id = str(task_id)
    for task in data['tasks']:
        if str(task['id']) == task_id:
            for key in ['content', 'date', 'status', 'source', 'source_dept', 'priority', 'notes', 'tags', 'project']:
                if key in req:
                    task[key] = req[key]
            # 同步 done 和 status
            if 'status' in task:
                task['done'] = (task['status'] == 'completed')
            elif 'done' in task:
                task['status'] = 'completed' if task['done'] else 'pending'
            break
    with open('data/daily_tasks.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return json_response({'task': task})


@app.route('/api/daily_tasks/<int:task_id>', methods=['DELETE'])
def delete_daily_task(task_id):
    with open('data/daily_tasks.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    task_id = str(task_id)
    data['tasks'] = [t for t in data['tasks'] if str(t['id']) != task_id]
    with open('data/daily_tasks.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return json_response({'success': True})


@app.route('/api/daily_tasks/<int:task_id>/toggle', methods=['POST'])
def toggle_daily_task(task_id):
    with open('data/daily_tasks.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    task_id = str(task_id)
    for task in data['tasks']:
        if str(task['id']) == task_id:
            task['done'] = not task.get('done', False)
            task['status'] = 'completed' if task['done'] else 'pending'
            task['completed_at'] = str(datetime.now()) if task['done'] else None
            break
    with open('data/daily_tasks.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return json_response({'task': task})


# ==================== 每日任务批量操作 API ====================

@app.route('/api/daily_tasks/batch_update', methods=['POST'])
def batch_update_daily_tasks():
    """批量更新任务日期：将原任务标记为延期，在新日期创建新任务"""
    req = request.get_json()
    task_ids = req.get('task_ids', [])
    new_date = req.get('new_date', '')

    if not task_ids:
        return jsonify({'error': 'task_ids 不能为空'}), 400
    if not new_date:
        return jsonify({'error': 'new_date 不能为空'}), 400
    try:
        datetime.strptime(new_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': '日期格式错误，应为 YYYY-MM-DD'}), 400

    with open('data/daily_tasks.json', 'r', encoding='utf-8') as f:
        data = json.load(f)

    moved = 0
    for task_id in task_ids:
        task_id_str = str(task_id)
        for task in data['tasks']:
            if str(task['id']) == task_id_str:
                # 原任务标记为延期
                task['status'] = 'delayed'
                task['original_date'] = task.get('date', '')

                # 创建新任务，继承原任务属性
                data['last_id'] += 1
                new_task = {
                    'id': data['last_id'],
                    'content': task.get('content', ''),
                    'date': new_date,
                    'status': 'pending',
                    'source': task.get('source', 'planned'),
                    'source_dept': task.get('source_dept', ''),
                    'priority': task.get('priority', 'medium'),
                    'notes': task.get('notes', ''),
                    'tags': task.get('tags', []),
                    'project': task.get('project', ''),
                    'created_at': str(datetime.now()),
                    'completed_at': None,
                    'done': False,
                    'original_task_id': task_id_str  # 关联原任务ID
                }
                data['tasks'].append(new_task)
                moved += 1
                break

    with open('data/daily_tasks.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    return jsonify({'success': True, 'moved': moved, 'new_date': new_date})


@app.route('/api/daily_tasks/batch_delete', methods=['POST'])
def batch_delete_daily_tasks():
    """批量删除任务"""
    req = request.get_json()
    task_ids = req.get('task_ids', [])

    if not task_ids:
        return jsonify({'error': 'task_ids 不能为空'}), 400

    with open('data/daily_tasks.json', 'r', encoding='utf-8') as f:
        data = json.load(f)

    task_ids_str = [str(tid) for tid in task_ids]
    original_count = len(data['tasks'])
    data['tasks'] = [t for t in data['tasks'] if str(t['id']) not in task_ids_str]
    deleted = original_count - len(data['tasks'])

    with open('data/daily_tasks.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    return jsonify({'success': True, 'deleted': deleted})


@app.route('/api/daily_tasks/dates_with_tasks', methods=['GET'])
def get_dates_with_tasks():
    """获取近14天有任务的日期列表"""
    today = datetime.now()
    days_ago_14 = today.replace(hour=0, minute=0, second=0, microsecond=0)

    with open('data/daily_tasks.json', 'r', encoding='utf-8') as f:
        data = json.load(f)

    # 统计每天的任务数量（排除已延期和已完成的）
    date_counts = {}
    for task in data.get('tasks', []):
        task_date = task.get('date', '')
        task_status = task.get('status', '')
        # 只统计未延期的任务
        if task_date and task_status != 'delayed':
            try:
                task_dt = datetime.strptime(task_date, '%Y-%m-%d')
                if task_dt >= days_ago_14:
                    date_counts[task_date] = date_counts.get(task_date, 0) + 1
            except ValueError:
                pass

    # 转换为列表并按日期排序
    result = [{'date': d, 'count': c} for d, c in date_counts.items()]
    result.sort(key=lambda x: x['date'], reverse=True)

    return jsonify(result)


# ==================== 项目管理 API（表单提交） ====================

@app.route('/add_company', methods=['POST'])
def add_company():
    """动态添加新公司"""
    company_name = request.form.get('company_name', '').strip()
    tab = request.form.get('tab', 'all')
    if not company_name:
        return redirect(url_for('index', tab=tab))
    data = load_data()
    # 避免重复添加
    if company_name not in data:
        data[company_name] = []
        save_data(data)
    return redirect(url_for('index', tab=tab))


@app.route('/get_companies', methods=['GET'])
def get_companies():
    """获取所有公司列表（用于动态下拉）"""
    data = load_data()
    companies = [g for g in data.keys() if g != 'meetings']
    return jsonify({'companies': companies})


@app.route('/add_project', methods=['POST'])
def add_project():
    """新建项目"""
    name = request.form.get('name', '').strip()
    group = request.form.get('group', '').strip()
    status = request.form.get('status', '进行中').strip()
    node = request.form.get('node', '').strip()
    deadline = request.form.get('deadline', '').strip()
    issues = request.form.get('issues', '').strip()
    tab = request.form.get('tab', 'all')

    # 立即添加的任务（可选）
    task_name = request.form.get('task_name', '').strip()
    task_priority = request.form.get('task_priority', 'medium').strip()
    task_due = request.form.get('task_due', '').strip()
    task_opinion = request.form.get('task_opinion', '').strip()

    if not name or not group:
        return redirect(url_for('index', tab=tab))

    data = load_data()
    if group not in data:
        data[group] = []
    if group == 'meetings':
        return redirect(url_for('index', tab=tab))

    # 生成项目ID（动态根据公司名生成）
    prefix = ''.join([c for c in group if c.isalnum()])[:8].lower()
    existing_ids = [p['id'] for p in data.get(group, []) if p['id'].startswith(prefix)]
    nums = [int(x.split('-')[-1]) for x in existing_ids if x.split('-')[-1].isdigit()]
    next_num = max(nums) + 1 if nums else 1
    project_id = f'{prefix}-{str(next_num).zfill(3)}'

    tasks = []
    if task_name:
        tasks.append({
            'text': html.escape(task_name),
            'done': False,
            'priority': task_priority if task_priority in ['low', 'medium', 'high'] else 'medium',
            'due': task_due,
            'opinion': html.escape(task_opinion),
            'createdAt': datetime.now().strftime('%Y-%m-%d')
        })

    # 计算 sort_order = 当前最大值+1
    existing_orders = [p.get('sort_order', 0) for p in data[group]]
    max_order = max(existing_orders) if existing_orders else -1
    new_sort_order = max_order + 1

    new_project = {
        'id': project_id,
        'name': html.escape(name),
        'status': status,
        'node': html.escape(node),
        'deadline': deadline,
        'issues': html.escape(issues),
        'tasks': tasks,
        'status_text': status,
        'status_color': '',
        'sort_order': new_sort_order
    }
    data[group].insert(0, new_project)
    save_data(data)
    return redirect(url_for('index', section='projects', tab=valid_tab(tab, data)) + f'#project-{project_id}')


@app.route('/delete_project', methods=['POST'])
def delete_project():
    """删除项目"""
    group = request.form.get('group', '').strip()
    project_id = request.form.get('project_id', '').strip()
    tab = request.form.get('tab', 'all')

    data = load_data()
    if group in data and group != 'meetings':
        data[group] = [p for p in data[group] if p.get('id') != project_id]

    save_data(data)
    return redirect(url_for('index', section='projects', tab=valid_tab(tab, data)))


@app.route('/add_task', methods=['POST'])
def add_task():
    """在项目下添加任务"""
    group = request.form.get('group', '').strip()
    project_id = request.form.get('project_id', '').strip()
    task_name = request.form.get('task_name', '').strip()
    priority = request.form.get('priority', 'medium').strip()
    due = request.form.get('due', '').strip()
    opinion = request.form.get('opinion', '').strip()
    createdAt = request.form.get('createdAt', '').strip()
    tab = request.form.get('tab', 'all')

    data = load_data()
    if not group or not project_id or not task_name:
        return redirect(url_for('index', tab=valid_tab(tab, data)) + f'#project-{project_id}')
    if group not in data or group == 'meetings':
        return redirect(url_for('index', tab=tab))

    for p in data[group]:
        if p.get('id') == project_id:
            p.setdefault('tasks', []).append({
                'text': html.escape(task_name),
                'done': False,
                'priority': priority if priority in ['low', 'medium', 'high'] else 'medium',
                'due': due,
                'opinion': html.escape(opinion),
                'createdAt': createdAt or datetime.now().strftime('%Y-%m-%d')
            })
            break

    save_data(data)
    return redirect(url_for('index', section='projects', tab=valid_tab(tab, data)) + f'#project-{project_id}')


@app.route('/delete_task', methods=['POST'])
def delete_task():
    """删除任务"""
    group = request.form.get('group', '').strip()
    project_id = request.form.get('project_id', '').strip()
    task_text = request.form.get('task_text', '').strip()
    tab = request.form.get('tab', 'all')

    data = load_data()
    if group in data and group != 'meetings':
        for p in data[group]:
            if p.get('id') == project_id:
                p['tasks'] = [t for t in p.get('tasks', []) if t.get('text') != task_text]
                break

    save_data(data)
    return redirect(url_for('index', section='projects', tab=valid_tab(tab, data)) + f'#project-{project_id}')


# ==================== 每日任务统计 API ====================

@app.route('/api/daily_tasks/stats', methods=['GET'])
def daily_tasks_stats():
    date_filter = request.args.get('date', '')
    with open('data/daily_tasks.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    tasks = data.get('tasks', [])
    if date_filter:
        tasks = [t for t in tasks if t.get('date') == date_filter]
    total = len(tasks)
    completed = len([t for t in tasks if t.get('status') == 'completed'])
    pending = len([t for t in tasks if t.get('status') == 'pending'])
    in_progress = len([t for t in tasks if t.get('status') == 'in_progress'])
    temp_count = len([t for t in tasks if t.get('source') == 'temp'])
    return json_response({
        'total': total,
        'completed': completed,
        'pending': pending,
        'in_progress': in_progress,
        'temp_count': temp_count,
        'completion_rate': round(completed / total * 100, 1) if total > 0 else 0
    })


# ==================== 导出功能 ====================

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def style_header(cell):
    """设置表头样式"""
    cell.font = Font(bold=True, color='FFFFFF', size=12)
    cell.fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='4B5563')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_data_cell(cell):
    """设置数据单元格样式"""
    cell.font = Font(size=11, color='3F3F3F')
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='374151')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def priority_label(p):
    if p == 'high': return '🔴高'
    if p == 'low': return '🟢低'
    return '🟡中'


@app.route('/export_projects', methods=['GET', 'POST'])
def export_projects():
    """导出项目任务 Excel，支持全量/按公司/按项目"""
    if not OPENPYXL_AVAILABLE:
        return jsonify({'error': 'openpyxl 未安装，请运行: pip install openpyxl'}), 500

    data = load_data()
    scope = request.values.get('scope', 'all')
    scope_key = request.values.get('key', '')

    # 准备数据
    rows = []
    header = ['公司', '项目', '任务名称', '优先级', '截止日期', '意见备注', '创建时间']

    companies = [g for g in data.keys() if g != 'meetings']

    if scope == 'company' and scope_key:
        # 按公司导出
        if scope_key in data:
            companies = [scope_key]
    elif scope == 'project' and scope_key:
        # 按项目导出：scope_key 格式为 "group|id"
        if '|' in scope_key:
            group, proj_id = scope_key.split('|', 1)
            for p in data.get(group, []):
                if p.get('id') == proj_id:
                    for t in p.get('tasks', []):
                        rows.append([
                            group,
                            p.get('name', ''),
                            t.get('text', ''),
                            priority_label(t.get('priority', 'medium')),
                            t.get('due', ''),
                            t.get('opinion', ''),
                            t.get('createdAt', ''),
                        ])
            companies = []  # 不再遍历所有公司

    for group in companies:
        for p in data.get(group, []):
            for t in p.get('tasks', []):
                rows.append([
                    group,
                    p.get('name', ''),
                    t.get('text', ''),
                    priority_label(t.get('priority', 'medium')),
                    t.get('due', ''),
                    t.get('opinion', ''),
                    t.get('createdAt', ''),
                ])

    # 生成 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = '项目任务'

    # 写入表头
    for col_idx, col_name in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        style_header(cell)

    # 写入数据
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            style_data_cell(cell)

    # 设置列宽
    col_widths = [15, 20, 35, 8, 12, 20, 12]
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

    # 生成文件名
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'项目任务_{ts}.xlsx'

    # 保存到内存并返回
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/export_daily', methods=['GET', 'POST'])
def export_daily():
    """导出每日任务 Excel，支持按日期范围筛选"""
    if not OPENPYXL_AVAILABLE:
        return jsonify({'error': 'openpyxl 未安装，请运行: pip install openpyxl'}), 500

    date_from = request.values.get('date_from', '')
    date_to = request.values.get('date_to', '')

    with open('data/daily_tasks.json', 'r', encoding='utf-8') as f:
        raw = json.load(f)
    tasks = raw.get('tasks', [])

    # 日期筛选
    if date_from:
        tasks = [t for t in tasks if t.get('date', '') >= date_from]
    if date_to:
        tasks = [t for t in tasks if t.get('date', '') <= date_to]

    # 排序：按日期升序
    tasks.sort(key=lambda t: t.get('date', ''))

    # 生成 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = '每日任务'

    header = ['日期', '任务内容', '完成状态', '创建时间']

    # 表头
    for col_idx, col_name in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        style_header(cell)

    # 数据
    for row_idx, t in enumerate(tasks, 2):
        done = t.get('done', False) or t.get('status') == 'completed'
        status_str = '✅已完成' if done else '☐未完成'
        # 兼容 content/text 字段
        content = t.get('content', '') or t.get('text', '')
        created = t.get('created_at', '')
        if created:
            created = created[:10] if len(created) > 10 else created

        row_data = [t.get('date', ''), content, status_str, created]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            style_data_cell(cell)

    # 列宽
    col_widths = [12, 40, 12, 12]
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'每日任务_{ts}.xlsx'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ==================== 启动 ====================

if __name__ == '__main__':
    print("=" * 50)
    print("📋 项目跟踪系统 v3.0")
    print("=" * 50)
    print("🌐 访问地址: http://localhost:8765")
    print("📝 按 Ctrl+C 停止服务")
    print("=" * 50)
    app.run(host='0.0.0.0', port=8765, debug=False)
